# import 2 package
import math
import openpyxl 

### SET DATA

# Name stocks

StocksName= ['HPG','VCB','VNM','BMP','GAS','IMP','NVL','SAB','MWG','MSN']

# Table1: Expected Return and Standard deviation
# row1: Expected Return
# row2: Standard deviation

table1 = [
    [2.31/100, 2.07/100, 0.03/100, 0.98/100, 1.65/100, 0.84/100, 0.26/100, 0.34/100, 1.04/100, 1.55/100],
    [12.14/100, 8.54/100, 7.02/100, 11.64/100, 11.03/100, 8.47/100, 14.67/100, 9.22/100, 11.28/100, 11.46/100]
]

# table 2: Covariance

table2=[
    [1.47/100,	0.53/100,	0.19/100,	0.44/100,	0.64/100,	0.45/100,	0.22/100,	0.31/100,	0.50/100,	0.57/100],
    [0.53/100,	0.73/100,	0.36/100,	0.34/100,	0.63/100,	0.19/100,	0.10/100,	0.31/100,	0.33/100,	0.28/100],
    [0.19/100,	0.36/100,	0.49/100,	0.30/100,	0.35/100,	0.07/100,	-0.12/100,	0.24/100,	0.28/100,	0.22/100],
    [0.44/100,	0.34/100,	0.30/100,	1.35/100,	0.36/100,	0.29/100,	0.12/100,	0.39/100,	0.50/100,	0.04/100],
    [0.64/100,	0.63/100,	0.35/100,	0.35/100,	1.22/100,	0.27/100,	0.15/100,	0.43/100,	0.55/100,	0.56/100],
    [0.45/100,	0.19/100,	0.07/100,	0.29/100,	0.27/100,	0.72/100,	-0.06/100,	0.29/100,	0.25/100,	0.13/100],
    [0.22/100,	0.10/100,	-0.12/100,	0.12/100,	0.15/100,	-0.06/100,	2.15/100,	-0.09/100,	0.30/100,	0.12/100],
    [0.31/100,	0.31/100,	0.24/100,	0.39/100,	0.43/100,	0.29/100,	-0.09/100,	0.85/100,	0.47/100,	0.18/100],
    [0.50/100,	0.33/100,	0.28/100,	0.50/100,	0.55/100,	0.25/100,	0.30/100,	0.47/100,	1.27/100,	0.34/100],
    [0.57/100,	0.28/100,	0.22/100,	0.04/100,	0.56/100,	0.13/100,	0.12/100,	0.18/100,	0.34/100,	1.31/100]
]	

# risk free rate
rf=(7/100)/12

# Asumming
A=20

# MaxSharp
MaxInforSharp=[]



# function: calculate sharp ratio
def sharpRatio(a,b):
    return (a-rf)/b



# function: find max sharp ratio , return max and weight
def findMaxSharp(E,O,Cov):
    # initialize weight wi , wj , wk
    w1=5.00  # wi
    w2=5.00  # wj
    w3=90 # wk
    
    arr=[]
    weight=[]
    cal= []
    while(w1<=90.00):
        i=w1
        j= w2
        while(j<=90):
            if i+j>95: 
                break
            k = 100-i-j

            a=i/100
            b=j/100
            c=k/100
          
            Ex= a*E[0] + b*E[1] + c*E[2]
            Op= math.sqrt(a**2 * O[0]**2 + b**2 * O[1]**2 + c**2 * O[2]**2 + 2*a*b*Cov[0] + 2*b*c*Cov[2] + 2*a*c*Cov[1])
            arr.append(sharpRatio(Ex,Op))                             # save value sharp
            weight.append((round(i,2),round(j,2),round(k,2)))         # save weight wi, wj, wk
            cal.append((round(Ex*100,2),round(Op*100,2)))             # save E(rp) and Op

            j+=1/4       
        w1+=1/4

    Max= max(arr)
    index= arr.index(Max)

    return Max, weight[index],cal[index]

# function: Optimal Risky Porfolio
def OptimalRiskyPorfolio(E_rp,op,y):
    U = y* E_rp + (1-y)*rf - 1/2*A* y**2 * op**2
    return U 

# function: write data on sheet "Tổ hợp"
def printSharpcombination():

    # open file
    wb = openpyxl.load_workbook('./Book1123.xlsx')
    sheet = wb['Tổ hợp']

    rowStart= 7
    count=1

    arr=[]
    # find max combination
    for i in range(0,10):
        for j in range(i+1,10):
            for k in range(j+1,10):
                E =[table1[0][i],table1[0][j],table1[0][k]]
                O =[table1[1][i],table1[1][j],table1[1][k]]
                Cov=[table2[i][j],table2[i][k],table2[j][k]]

                result= findMaxSharp(E,O,Cov)
                arr.append(result[0])

    MaxSharp = max(arr)

    # write data
    for i in range(0,10):
        for j in range(i+1,10):
            for k in range(j+1,10):
                E =[table1[0][i],table1[0][j],table1[0][k]]
                O =[table1[1][i],table1[1][j],table1[1][k]]
                Cov=[table2[i][j],table2[i][k],table2[j][k]]

                result= findMaxSharp(E,O,Cov)
                print('No:',count,'- Name:', StocksName[i],StocksName[j],StocksName[k])
                print(F"{round(result[0]*100,2)}%",result[1])
                print("-----------------------")

                sheet.cell(row=rowStart+count, column=1, value=count)
                if result[0] == MaxSharp:
                    sheet.cell(row=rowStart+count, column=2, value=f"{StocksName[i]}-{StocksName[j]}-{StocksName[k]} (Max)")
                    # get data row Max Sharp
                    MaxInforSharp.append((StocksName[i],StocksName[j],StocksName[k])) # name
                    MaxInforSharp.append(result[1])                                   # weight wi, wj, wk
                    MaxInforSharp.append((E[0]*100,E[1]*100,E[2]*100))                # Er_i, E_j, E_k 
                    MaxInforSharp.append((O[0]*100,O[1]*100,O[2]*100))                # O_i, O_j, O_k
                    MaxInforSharp.append((Cov[0]*100,Cov[1]*100,Cov[2]*100))          # covariance 
                    MaxInforSharp.append(result[2])                                   # E_rp, op   
                    MaxInforSharp.append(result[0])                                   # Sharp ratio

                else:
                    sheet.cell(row=rowStart+count, column=2, value=f"{StocksName[i]}-{StocksName[j]}-{StocksName[k]}")
                
                sheet.cell(row=rowStart+count, column=3, value=f"{StocksName[i]}")
                sheet.cell(row=rowStart+count, column=4, value=f"{StocksName[j]}")
                sheet.cell(row=rowStart+count, column=5, value=f"{StocksName[k]}")
                sheet.cell(row=rowStart+count, column=6, value=f"{result[1][0]}%")
                sheet.cell(row=rowStart+count, column=7, value=f"{result[1][1]}%")
                sheet.cell(row=rowStart+count, column=8, value=f"{result[1][2]}%")
                sheet.cell(row=rowStart+count, column=9, value=f"100%")
                sheet.cell(row=rowStart+count, column=10, value=f"{E[0]*100}%")
                sheet.cell(row=rowStart+count, column=11, value=f"{E[1]*100}%")
                sheet.cell(row=rowStart+count, column=12, value=f"{E[2]*100}%")
                sheet.cell(row=rowStart+count, column=13, value=f"{O[0]*100}%")
                sheet.cell(row=rowStart+count, column=14, value=f"{O[1]*100}%")
                sheet.cell(row=rowStart+count, column=15, value=f"{O[2]*100}%")
                sheet.cell(row=rowStart+count, column=16, value=f"{Cov[0]*100}%")
                sheet.cell(row=rowStart+count, column=17, value=f"{Cov[1]*100}%")
                sheet.cell(row=rowStart+count, column=18, value=f"{Cov[2]*100}%")
                sheet.cell(row=rowStart+count, column=19, value=f"{result[2][0]}%")
                sheet.cell(row=rowStart+count, column=20, value=f"{result[2][1]}%")
                sheet.cell(row=rowStart+count, column=21, value=f"{round(result[0]*100,2)}%")

                count+=1
               
    # save file and close
    wb.save('./Book1123.xlsx')



# function: write data on sheet "TheOptimalRiskyPorfolio"
def printOptimalRiskyPorfolio():

    # open file
    wb = openpyxl.load_workbook('./Book1123.xlsx')
    sheet = wb['TheOptimalRiskyPorfolio']


    i=0.00
    E_rp=MaxInforSharp[5][0]
    op=MaxInforSharp[5][1]
    arr=[]
    weight=[]
    while(i<=100):
        y=i
        arr.append(OptimalRiskyPorfolio(E_rp/100,op/100,y/100))
        weight.append(y)
        i+=0.01

    Max= max(arr)
    index= arr.index(Max)

    # table1
    sheet.cell(row=6, column=3, value=f"{MaxInforSharp[2][0]}%")
    sheet.cell(row=7, column=3, value=f"{MaxInforSharp[2][1]}%")
    sheet.cell(row=8, column=3, value=f"{MaxInforSharp[2][2]}%")
    sheet.cell(row=9, column=3, value=f"{MaxInforSharp[4][0]}%")
    sheet.cell(row=10, column=3, value=f"{MaxInforSharp[4][1]}%")
    sheet.cell(row=11, column=3, value=f"{MaxInforSharp[4][2]}%")
    sheet.cell(row=6, column=4, value=f"{MaxInforSharp[4][0]}%")
    sheet.cell(row=7, column=4, value=f"{MaxInforSharp[4][1]}%")
    sheet.cell(row=8, column=4, value=f"{MaxInforSharp[4][2]}%")

    #table2
    sheet.cell(row=4, column=7, value=f"{MaxInforSharp[1][0]}%")
    sheet.cell(row=5, column=7, value=f"{MaxInforSharp[1][1]}%")
    sheet.cell(row=6, column=7, value=f"{MaxInforSharp[1][2]}%")
    sheet.cell(row=9, column=7, value=f"{MaxInforSharp[5][1]}%")
    sheet.cell(row=10, column=7, value=f"{round(MaxInforSharp[6]*100,2)}%")

    #table 3
    sheet.cell(row=15, column=3, value=f"{round(weight[index],2)}%")

    # print result on terminal

    print("Max sharpe ratio:",round(MaxInforSharp[6]*100,2))
    print("OptimalRiskyPorfolio:",round(max(arr),5),'- y:',round(weight[index],2))


    # save file and close
    wb.save('./Book1123.xlsx')




printSharpcombination()
printOptimalRiskyPorfolio()

                







    